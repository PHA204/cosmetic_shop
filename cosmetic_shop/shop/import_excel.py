from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import HttpResponse
import pandas as pd
from openpyxl import Workbook
import openpyxl
from io import BytesIO
from .models import Product, Category

# ═══════════════════════════════════════════════════════════════════════════════════
# IMPORT TỪ EXCEL
# ═══════════════════════════════════════════════════════════════════════════════════

@login_required
@user_passes_test(lambda u: u.is_staff)
def admin_import_products(request):
    """Trang upload file Excel để nhập sản phẩm"""
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, 'Vui lòng chọn file Excel.')
            return render(request, 'shop/admin/import_products.html', {
                'categories': Category.objects.all(),
            })

        # Kiểm tra đuôi file
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'Vui lòng chọn file Excel (.xlsx hoặc .xls)')
            return render(request, 'shop/admin/import_products.html', {
                'categories': Category.objects.all(),
            })

        try:
            # Đọc file Excel
            df = pd.read_excel(excel_file)
            
            # Chuẩn hóa tên cột (lower, strip, thay space thành _)
            df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')
            
            # Các cột bắt buộc
            required_cols = ['name', 'price']
            missing_cols = [col for col in required_cols if col not in df.columns]
            if missing_cols:
                messages.error(request, f'Thiếu cột bắt buộc: {missing_cols}. '
                    f'Các cột cần thiết: name, price (có thể có: description, stock, category)')
                return render(request, 'shop/admin/import_products.html', {
                    'categories': Category.objects.all(),
                })

            # Map tên category -> id
            category_map = {c.name.lower(): c.id for c in Category.objects.all()}
            category_map.update({c.slug: c.id for c in Category.objects.all()})
            
            created = 0
            updated = 0
            errors = []
            
            for index, row in df.iterrows():
                try:
                    name = str(row.get('name', '')).strip()
                    if not name:
                        errors.append(f'Dòng {index + 2}: Thiếu tên sản phẩm')
                        continue

                    # Xử lý giá
                    try:
                        price = int(float(row.get('price', 0)))
                    except (ValueError, TypeError):
                        errors.append(f'Dòng {index + 2}: Giá không hợp lệ')
                        continue

                    # Xử lý tồn kho
                    stock = 0
                    if 'stock' in df.columns:
                        try:
                            stock = int(float(row.get('stock', 0)))
                        except (ValueError, TypeError):
                            stock = 0

                    # Xử lý mô tả
                    description = ''
                    if 'description' in df.columns:
                        description = str(row.get('description', ''))

                    # Xử lý category
                    category_id = None
                    if 'category' in df.columns:
                        cat_name = str(row.get('category', '')).strip().lower()
                        if cat_name and cat_name in category_map:
                            category_id = category_map[cat_name]

                    # Kiểm tra sản phẩm đã tồn tại chưa
                    product = Product.objects.filter(name__iexact=name).first()
                    if product:
                        # Cập nhật
                        product.price = price
                        product.stock = stock
                        product.description = description or product.description
                        if category_id:
                            product.category_id = category_id
                        product.save()
                        updated += 1
                    else:
                        # Tạo mới
                        Product.objects.create(
                            name=name,
                            price=price,
                            stock=stock,
                            description=description,
                            category_id=category_id,
                        )
                        created += 1

                except Exception as e:
                    errors.append(f'Dòng {index + 2}: {str(e)}')

            # Kết quả
            if created > 0:
                messages.success(request, f'Đã tạo {created} sản phẩm mới.')
            if updated > 0:
                messages.success(request, f'Đã cập nhật {updated} sản phẩm.')
            if errors:
                for err in errors[:10]:  # Chỉ hiện 10 lỗi đầu
                    messages.warning(request, err)
            
            return redirect('shop:admin_products')

        except Exception as e:
            messages.error(request, f'Lỗi khi đọc file: {str(e)}')
            return render(request, 'shop/admin/import_products.html', {
                'categories': Category.objects.all(),
            })

    return render(request, 'shop/admin/import_products.html', {
        'categories': Category.objects.all(),
    })


@login_required
@user_passes_test(lambda u: u.is_staff)
def admin_download_template(request):
    """Tải file mẫu Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    
    # Header
    ws.append(["name", "price", "description", "stock", "category"])
    
    # Ví dụ
    ws.append(["Son Môi Đỏ", 150000, "Son môi đỏ cherry", 50, "Son môi"])
    ws.append(["Kem Dưỡng Ẩm", 200000, "Kem dưỡng ẩm cho da", 30, "Kem dưỡng"])
    
    # Format header
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="CCCCCC")
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Response
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="san_pham_template.xlsx"'
    return response